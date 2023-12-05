import json
import random
import time
import datetime
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange

POSSIBLE_FORMATS = """
    The possible formats of the power timeseries are :
        - a single value (int or float) if the power is constant throughout the load_profile (with random fluctuations if the variable 'thermal_p_var' is provided)
        - an array of value provided as text in a json array format, i.e. : [val1, val2, val3, ...]
        - a range of cells in another sheet of the input file (type '=' in the cell and then select the wished range of values to get the correct format automatically)

    ***Note***
    The last two formats will only accept one column (values only) or two columns (timestamps and values, respectively) and exactly 365 rows
"""


def read_input_file(filename):
    """Parse a RAMP .xlsx input file"""

    wb = load_workbook(filename=filename)
    sheet_names = wb.sheetnames
    name = sheet_names[0]
    headers = [c.value for c in wb[name][1]]
    df = pd.DataFrame(tuple(wb[name].values)[1:], columns=headers)
    df = df.fillna(value=np.nan)

    df["p_series"] = False
    for i, v in enumerate(df["power"].values):
        if isinstance(v, str):
            appliance_name = df["name"].iloc[i]
            user_name = df["user_name"].iloc[i]
            # the timeseries is provided as a range of values in the spreadsheet
            if "=" in v:
                ts_sheet_name, ts_range = v.replace("=", "").split("!")
                cr = CellRange(ts_range)
                if cr.size["columns"] == 1:
                    ts = json.dumps(
                        [
                            wb[ts_sheet_name].cell(row=c[0], column=c[1]).value
                            for c in cr.cells
                        ]
                    )
                elif cr.size["columns"] == 2:
                    ts = pd.DataFrame(
                        [
                            [
                                wb[ts_sheet_name].cell(row=c[0], column=c[1]).value
                                for c in col
                            ]
                            for col in cr.cols
                        ]
                    ).T
                    ts = ts.to_json(orient="values")
                else:
                    raise (
                        ValueError(
                            f"The provided range for the power timeseries of the appliance '{appliance_name}' of user '{user_name} spans more than two columns in '{filename}' (range {ts_range} of sheet '{ts_sheet_name}')\n{POSSIBLE_FORMATS}"
                        )
                    )
                if cr.size["rows"] != 366:
                    raise (
                        ValueError(
                            f"The provided range for the power timeseries of the appliance '{appliance_name}' of user '{user_name}' in '{filename}' does not contain 366 values as expected  (range {ts_range} of sheet '{ts_sheet_name}')\n{POSSIBLE_FORMATS}"
                        )
                    )
            # the timeseries is expected as an array in json format
            else:
                try:
                    ts = json.loads(v)
                    if len(ts) != 366:
                        raise (
                            ValueError(
                                f"The provided power timeseries of the appliance '{appliance_name}' of user '{user_name}' in '{filename}' does not contain 366 values as expected\n{POSSIBLE_FORMATS}"
                            )
                        )
                    ts = v
                except json.JSONDecodeError:
                    raise (
                        ValueError(
                            f"Could not parse the power timeseries provided for appliance '{appliance_name}' of user '{user_name}' in '{filename}'\n{POSSIBLE_FORMATS}"
                        )
                    )
            df.loc[i, "power"] = ts
            df.loc[i, "p_series"] = True
    return df


def random_variation(var, norm=1):
    """Pick a random variable within a uniform distribution of range [1-var, 1+var]

    Parameters
    ----------
    var: float
        sets the range of the uniform distribution around one
    norm: float
        multiplication factor of the random variable, default = 1

    Returns
    -------
    random number close to norm
    """
    return norm * random.uniform((1 - var), (1 + var))


def duty_cycle(var, t1, p1, t2, p2):
    """Assign a two period duty cycle

    concatenate an array where values equal p1 for a time (t1 +- random variation)
    followed by values equal to p2 for a time (t2 +- random variation)

    Parameters
    ----------
    var: float
        sets the range of the uniform distribution around t1 and t2
    t1: int
        time interval of the first part of the duty cycle in minutes
    p1: float
        power of the first part of the duty cycle in Watt
    t2: int
        time interval of the second part of the duty cycle in minutes
    p2: int
        power of the second part of the duty cycle in Watt

    Returns
    -------
    Power during each timestep of the duty cycle where p1 is repeated (t1 +- random variation) times and p2 is repeated (t2 +- random variation) times.
    The duty cycle is implicitly sampled every minutes (which is the unit for t1 and t2)
    """
    return np.concatenate(
        (
            np.ones(int(random_variation(var=-var, norm=t1))) * p1,
            np.ones(int(random_variation(var=-var, norm=t2))) * p2,
        )
    )


def random_choice(var, t1, p1, t2, p2):
    """Chooses one of two duty cycles randomly

    The choice is between a normal duty cycle and a reversed duty cycle (where t1 is swapped with t2 and p1 with p2)

    Parameters
    ----------
    var: float
        sets the range of the uniform distribution around t1 and t2
    t1: int
        time interval of the first part of the duty cycle in minutes
    p1: float
        power of the first part of the duty cycle in Watt
    t2: int
        time interval of the second part of the duty cycle in minutes
    p2: int
        power of the second part of the duty cycle in Watt

    Returns
    -------
    A duty cycle, see function duty_cycle
    """
    return random.choice(
        [
            duty_cycle(var, t1=t1, p1=p1, t2=t2, p2=p2),
            duty_cycle(var, t1=t2, p1=p2, t2=t1, p2=p1),
        ]
    )


def get_day_type(day):
    """Given a datetime object return 0 for weekdays or 1 for weekends"""

    if isinstance(day, str):
        day = datetime.date.fromisoformat(day)

    if day.weekday() > 4:
        answer = 1
    else:
        answer = 0
    return answer


def yearly_pattern(year=None):
    """
    Definition of a yearly pattern of weekends and weekdays, in case some appliances have specific wd/we behaviour
    If no argument is provided, the pattern always starts a monday and lasts 365 days, otherwise the pattern matches
    the weekdays and weekends of the provided year
    """
    if year is None:
        year_behaviour = np.zeros(365)
        year_behaviour[5:365:7] = 1
        year_behaviour[6:365:7] = 1
        year_behaviour = year_behaviour.tolist()
    else:
        # a list with 0 for weekdays and 1 for weekends
        year_behaviour = (
            pd.date_range(start=f"{year}-01-01", end=f"{year}-12-31", freq="D")
            .map(get_day_type)
            .to_list()
        )
    return year_behaviour


def within_peak_time_window(win_start, win_stop, peak_win_start, peak_win_stop):
    """Given determines if a switch on window falls within the peak time window"""
    answer = True
    # start and stop of the given window are both below the lower limit of peak time window
    if win_start < peak_win_start and win_stop < peak_win_start:
        answer = False
    # start and stop of the given window are both above the upper limit of peak time window
    if win_start > peak_win_stop and win_stop > peak_win_stop:
        answer = False
    return answer


def calc_time_taken(func):
    """Calculates the time elapsed during the execution of a function"""

    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        end = time.time()
        print(
            func.__name__
            + " required "
            + str((end - start) * 1)
            + " seconds for execution. "
        )
        return result

    return wrapper
